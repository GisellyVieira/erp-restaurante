from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from sqlalchemy import inspect, text

from models import (
    db,
    Usuario,
    Insumo,
    Produto,
    Venda,
    MovimentacaoEstoque,
    MovimentacaoProduto,
    FichaTecnica,
    Financeiro,
)

import io
import os
import shutil
from datetime import datetime
import traceback


app = Flask(__name__)

app.config["SECRET_KEY"] = os.environ.get(
    "SECRET_KEY",
    "chave-local-altere-em-producao",
)

app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL",
    "sqlite:///database.db",
)

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,
    "pool_recycle": 300,
}

db.init_app(app)


def criar_banco():
    """Cria as tabelas e acrescenta colunas novas em bancos já existentes."""

    db.create_all()

    inspector = inspect(db.engine)
    tabelas = inspector.get_table_names()

    # ==================================================
    # AJUSTES NA TABELA FICHA_TECNICA
    # ==================================================

    if "ficha_tecnica" in tabelas:
        colunas_ficha = {
            coluna["name"]
            for coluna in inspector.get_columns(
                "ficha_tecnica"
            )
        }

        if "produto_base_id" not in colunas_ficha:
            with db.engine.begin() as conexao:
                conexao.execute(
                    text(
                        """
                        ALTER TABLE ficha_tecnica
                        ADD COLUMN produto_base_id INTEGER
                        REFERENCES produto(id)
                        """
                    )
                )

            print(
                "Coluna produto_base_id criada com sucesso."
            )

    # ==================================================
    # AJUSTES NA TABELA PRODUTO
    # ==================================================

    if "produto" in tabelas:
        colunas_produto = {
            coluna["name"]
            for coluna in inspector.get_columns(
                "produto"
            )
        }

        if "finalidade" not in colunas_produto:
            with db.engine.begin() as conexao:
                conexao.execute(
                    text(
                        """
                        ALTER TABLE produto
                        ADD COLUMN finalidade VARCHAR(30)
                        NOT NULL DEFAULT 'Venda'
                        """
                    )
                )

            print(
                "Coluna finalidade criada com sucesso."
            )

        if (
            "rendimento_quantidade"
            not in colunas_produto
        ):
            with db.engine.begin() as conexao:
                conexao.execute(
                    text(
                        """
                        ALTER TABLE produto
                        ADD COLUMN rendimento_quantidade
                        FLOAT DEFAULT 1
                        """
                    )
                )

            print(
                "Coluna rendimento_quantidade "
                "criada com sucesso."
            )

        if "rendimento_unidade" not in colunas_produto:
            with db.engine.begin() as conexao:
                conexao.execute(
                    text(
                        """
                        ALTER TABLE produto
                        ADD COLUMN rendimento_unidade
                        VARCHAR(20) DEFAULT 'un'
                        """
                    )
                )

            print(
                "Coluna rendimento_unidade "
                "criada com sucesso."
            )
                # ==================================================
    # AJUSTES NA TABELA INSUMO
    # ==================================================

    if "insumo" in tabelas:
        colunas_insumo = {
            coluna["name"]
            for coluna in inspector.get_columns(
                "insumo"
            )
        }

        if "demanda_mensal_estimada" not in colunas_insumo:
            with db.engine.begin() as conexao:
                conexao.execute(
                    text(
                        """
                        ALTER TABLE insumo
                        ADD COLUMN demanda_mensal_estimada
                        FLOAT NOT NULL DEFAULT 0
                        """
                    )
                )

            print(
                "Coluna demanda_mensal_estimada criada com sucesso."
            )

        if "custo_pedido" not in colunas_insumo:
            with db.engine.begin() as conexao:
                conexao.execute(
                    text(
                        """
                        ALTER TABLE insumo
                        ADD COLUMN custo_pedido
                        FLOAT NOT NULL DEFAULT 0
                        """
                    )
                )

            print(
                "Coluna custo_pedido criada com sucesso."
            )

        if "percentual_armazenagem" not in colunas_insumo:
            with db.engine.begin() as conexao:
                conexao.execute(
                    text(
                        """
                        ALTER TABLE insumo
                        ADD COLUMN percentual_armazenagem
                        FLOAT NOT NULL DEFAULT 10
                        """
                    )
                )

            print(
                "Coluna percentual_armazenagem criada com sucesso."
            )
            # ==================================================
# AJUSTES NA TABELA VENDA
# ==================================================

    if "venda" in tabelas:
        colunas_venda = {
        coluna["name"]
        for coluna in inspector.get_columns("venda")
    }

    if "movimentou_estoque" not in colunas_venda:
        with db.engine.begin() as conexao:
            conexao.execute(
                text(
                    """
                    ALTER TABLE venda
                    ADD COLUMN movimentou_estoque
                    BOOLEAN NOT NULL DEFAULT TRUE
                    """
                )
            )

        print(
            "Coluna movimentou_estoque criada com sucesso."
        )

def usuario_logado():
    return "usuario_id" in session


def converter_float(valor, padrao=0.0):
    """Converte campos numéricos aceitando ponto ou vírgula."""
    if valor is None:
        return padrao

    texto_valor = str(valor).strip()

    if not texto_valor:
        return padrao

    try:
        return float(texto_valor.replace(",", "."))
    except ValueError:
        return padrao


def ficha_cria_ciclo(produto_id, produto_base_id, visitados=None):
    """Verifica se a inclusão de um produto-base criaria referência circular."""
    if produto_id == produto_base_id:
        return True

    if visitados is None:
        visitados = set()

    if produto_base_id in visitados:
        return False

    visitados.add(produto_base_id)

    itens_base = FichaTecnica.query.filter_by(
        produto_id=produto_base_id
    ).all()

    for item in itens_base:
        if item.produto_base_id:
            if item.produto_base_id == produto_id:
                return True

            if ficha_cria_ciclo(
                produto_id,
                item.produto_base_id,
                visitados,
            ):
                return True

    return False


def calcular_consumo_insumos(produto, multiplicador=1.0, caminho=None):
    """
    Expande a ficha técnica do produto e retorna uma lista de consumos:
    [(insumo, quantidade_para_baixa), ...].

    Produtos-base são abertos recursivamente até chegar aos insumos.
    """
    if caminho is None:
        caminho = set()

    if produto.id in caminho:
        raise ValueError(
            f"Foi encontrada uma referência circular na ficha de {produto.nome}."
        )

    caminho_atual = set(caminho)
    caminho_atual.add(produto.id)

    consumos = []

    for item in produto.ficha_itens:
        quantidade_item = (
            item.quantidade_convertida_para_estoque()
            * multiplicador
        )

        if item.insumo_id:
            if item.insumo is None:
                raise ValueError(
                    f"Há um insumo inválido na ficha de {produto.nome}."
                )

            consumos.append((item.insumo, quantidade_item))
            continue

        if item.produto_base_id:
            produto_base = Produto.query.get(item.produto_base_id)

            if produto_base is None:
                raise ValueError(
                    f"Há um produto-base inválido na ficha de {produto.nome}."
                )

            consumos.extend(
                calcular_consumo_insumos(
                    produto_base,
                    multiplicador=quantidade_item,
                    caminho=caminho_atual,
                )
            )
            continue

        raise ValueError(
            f"Há um item sem insumo ou produto-base na ficha de {produto.nome}."
        )

    return consumos


with app.app_context():
    criar_banco()


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        nome_usuario = request.form.get("usuario", "").strip()
        senha = request.form.get("senha", "")

        user = Usuario.query.filter_by(
            usuario=nome_usuario
        ).first()

        if user and user.verificar_senha(senha):
            session["usuario_id"] = user.id
            session["usuario_nome"] = user.nome
            return redirect(url_for("dashboard"))

        flash("Usuário ou senha incorretos!")

    return render_template("login.html")


@app.route("/dashboard")
def dashboard():
    if not usuario_logado():
        return redirect(url_for("login"))

    insumos_lista = Insumo.query.all()
    produtos_lista = Produto.query.all()
    vendas_lista = Venda.query.all()
    financeiros = Financeiro.query.all()

    receita_vendas = sum(
        venda.receita_total or 0
        for venda in vendas_lista
    )

    cmv_total = sum(
        venda.cmv_total or 0
        for venda in vendas_lista
    )

    margem_total = sum(
        venda.margem_total or 0
        for venda in vendas_lista
    )

    entradas_financeiras = sum(
        financeiro.valor or 0
        for financeiro in financeiros
        if financeiro.tipo == "Entrada"
    )

    despesas_operacionais = sum(
        financeiro.valor or 0
        for financeiro in financeiros
        if financeiro.tipo == "Saída"
    )

    receita_total = receita_vendas + entradas_financeiras
    lucro_operacional = margem_total - despesas_operacionais

    margem_percentual = 0

    if receita_vendas > 0:
        margem_percentual = (
            margem_total / receita_vendas
        ) * 100

    valor_estoque = sum(
        insumo.estoque_atual()
        * insumo.custo_medio_unitario()
        for insumo in insumos_lista
    )

    itens_ponto_pedido = sum(
        1
        for insumo in insumos_lista
        if insumo.status_estoque() == "Ponto de pedido"
    )

    cobertura_baixa = sum(
        1
        for insumo in insumos_lista
        if insumo.status_estoque() == "Cobertura baixa"
    )

    insumos_sem_estoque = [
        insumo
        for insumo in insumos_lista
        if insumo.estoque_atual() <= 0
    ]

    insumos_ponto_pedido = [
        insumo
        for insumo in insumos_lista
        if insumo.status_estoque() == "Ponto de pedido"
    ]

    insumos_cobertura_baixa = [
        insumo
        for insumo in insumos_lista
        if insumo.status_estoque() == "Cobertura baixa"
    ]

    produtos_margem_baixa = [
        produto
        for produto in produtos_lista
        if produto.percentual_margem() < 40
    ]

    return render_template(
        "dashboard.html",
        nome=session.get("usuario_nome"),
        total_insumos=len(insumos_lista),
        total_produtos=len(produtos_lista),
        produtos_ativos=sum(
            1
            for produto in produtos_lista
            if produto.ativo
        ),
        valor_estoque=valor_estoque,
        itens_ponto_pedido=itens_ponto_pedido,
        cobertura_baixa=cobertura_baixa,
        receita_total=receita_total,
        receita_vendas=receita_vendas,
        cmv_total=cmv_total,
        margem_total=margem_total,
        margem_percentual=margem_percentual,
        despesas_operacionais=despesas_operacionais,
        lucro_operacional=lucro_operacional,
        insumos_sem_estoque=insumos_sem_estoque,
        insumos_ponto_pedido=insumos_ponto_pedido,
        insumos_cobertura_baixa=insumos_cobertura_baixa,
        produtos_margem_baixa=produtos_margem_baixa,
    )


@app.route("/insumos", methods=["GET", "POST"])
def insumos():
    if not usuario_logado():
        return redirect(url_for("login"))

    if request.method == "POST":

        nome = request.form.get(
            "nome",
            ""
        ).strip()

        unidade = request.form.get(
            "unidade",
            ""
        ).strip()

        categoria = request.form.get(
            "categoria",
            "Matéria-prima"
        ).strip()

        if not nome:
            flash(
                "Informe o nome do insumo."
            )
            return redirect(url_for("insumos"))

        if not unidade:
            flash(
                "Selecione a unidade de medida."
            )
            return redirect(url_for("insumos"))

        try:

            novo = Insumo(
                nome=nome,
                unidade=unidade,
                categoria=categoria or "Matéria-prima"
            )

            db.session.add(novo)
            db.session.commit()

            flash(
                "Insumo cadastrado com sucesso!"
            )

        except Exception as erro:

            db.session.rollback()

            print(
                erro,
                flush=True
            )

            flash(
                "Não foi possível cadastrar o insumo."
            )

        return redirect(
            url_for("insumos")
        )

    lista = Insumo.query.order_by(
        Insumo.nome.asc()
    ).all()

    return render_template(
        "insumos.html",
        insumos=lista
    )


@app.route(
    "/entrada_estoque/<int:insumo_id>",
    methods=["POST"]
)
def entrada_estoque(insumo_id):
    if not usuario_logado():
        return redirect(url_for("login"))

    insumo = Insumo.query.get_or_404(insumo_id)

    quantidade = converter_float(
        request.form.get("quantidade")
    )

    valor_total = converter_float(
        request.form.get("valor_total")
    )

    if quantidade <= 0:
        flash(
            "A quantidade da entrada deve ser maior que zero."
        )
        return redirect(url_for("insumos"))

    if valor_total < 0:
        flash("O valor total não pode ser negativo.")
        return redirect(url_for("insumos"))

    try:
        entrada = MovimentacaoEstoque(
            insumo_id=insumo.id,
            tipo="Entrada",
            quantidade=quantidade,
            valor_total=valor_total,
            observacao="Compra registrada",
        )

        saida_financeira = Financeiro(
            tipo="Saída",
            categoria="Compra de insumos",
            descricao=f"Compra de {insumo.nome}",
            valor=valor_total,
        )

        db.session.add(entrada)
        db.session.add(saida_financeira)
        db.session.commit()

        flash(
            "Entrada de estoque registrada com sucesso!"
        )

    except Exception:
        db.session.rollback()

        flash(
            "Não foi possível registrar a entrada de estoque."
        )

    return redirect(url_for("insumos"))


@app.route(
    "/insumos/editar/<int:id>",
    methods=["GET", "POST"]
)
def editar_insumo(id):

    if not usuario_logado():
        return redirect(url_for("login"))

    insumo = Insumo.query.get_or_404(id)

    if request.method == "POST":

        nome = request.form.get(
            "nome",
            ""
        ).strip()

        unidade = request.form.get(
            "unidade",
            ""
        ).strip()

        categoria = request.form.get(
            "categoria",
            "Matéria-prima"
        ).strip()

        if not nome:

            flash(
                "Informe o nome do insumo."
            )

            return redirect(
                url_for(
                    "editar_insumo",
                    id=id
                )
            )

        if not unidade:

            flash(
                "Selecione a unidade."
            )

            return redirect(
                url_for(
                    "editar_insumo",
                    id=id
                )
            )

        try:

            insumo.nome = nome
            insumo.unidade = unidade
            insumo.categoria = categoria

            db.session.commit()

            flash(
                "Insumo atualizado com sucesso!"
            )

        except Exception as erro:

            db.session.rollback()

            print(
                erro,
                flush=True
            )

            flash(
                "Não foi possível atualizar o insumo."
            )

        return redirect(
            url_for("insumos")
        )

    return render_template(
        "editar_insumo.html",
        insumo=insumo
    )

@app.route(
    "/excluir_insumo/<int:id>",
    methods=["POST", "GET"]
)
def excluir_insumo(id):
    if not usuario_logado():
        return redirect(url_for("login"))

    insumo = Insumo.query.get_or_404(id)

    if getattr(insumo, "movimentacoes", None):
        flash(
            "Este insumo não pode ser excluído "
            "porque possui movimentações."
        )

        return redirect(url_for("insumos"))

    itens_ficha = FichaTecnica.query.filter_by(
        insumo_id=insumo.id
    ).count()

    if itens_ficha:
        flash(
            "Este insumo não pode ser excluído "
            "porque está em uma ficha técnica."
        )

        return redirect(url_for("insumos"))

    try:
        db.session.delete(insumo)
        db.session.commit()

        flash("Insumo excluído com sucesso!")

    except Exception:
        db.session.rollback()

        flash(
            "Não foi possível excluir o insumo."
        )

    return redirect(url_for("insumos"))

# =========================================================
# FUNÇÃO AUXILIAR PARA VALORES DO FORMULÁRIO
# =========================================================

def converter_numero_formulario(
    valor,
    padrao=0
):
    """
    Converte valores recebidos do formulário para float.

    Aceita:
    10
    10.50
    10,50
    """

    if valor is None:
        return float(padrao)

    texto = str(valor).strip()

    if not texto:
        return float(padrao)

    texto = texto.replace(" ", "")

    if "," in texto and "." in texto:
        texto = texto.replace(".", "")
        texto = texto.replace(",", ".")

    elif "," in texto:
        texto = texto.replace(",", ".")

    try:
        return float(texto)

    except (TypeError, ValueError):
        return float(padrao)


# =========================================================
# PRODUTOS — LISTAGEM E CADASTRO
# =========================================================

@app.route(
    "/produtos",
    methods=["GET", "POST"]
)
def produtos():
    if "usuario_id" not in session:
        return redirect(
            url_for("login")
        )

    if request.method == "POST":
        nome = (
            request.form.get("nome", "")
            .strip()
        )

        categoria = (
            request.form.get(
                "categoria",
                "Outros"
            )
            .strip()
        )

        tipo_produto = (
            request.form.get(
                "tipo_produto",
                "Produzido"
            )
            .strip()
        )

        finalidade = (
            request.form.get(
                "finalidade",
                "Venda"
            )
            .strip()
        )

        preco_venda = converter_numero_formulario(
            request.form.get("preco_venda"),
            0
        )

        custo_compra = converter_numero_formulario(
            request.form.get("custo_compra"),
            0
        )

        estoque_produto = converter_numero_formulario(
            request.form.get("estoque_produto"),
            0
        )

        rendimento_quantidade = (
            converter_numero_formulario(
                request.form.get(
                    "rendimento_quantidade"
                ),
                1
            )
        )

        rendimento_unidade = (
            request.form.get(
                "rendimento_unidade",
                "un"
            )
            .strip()
        )

        if not nome:
            flash(
                "Informe o nome do produto.",
                "erro"
            )

            return redirect(
                url_for("produtos")
            )

        produto_existente = (
            Produto.query
            .filter(
                db.func.lower(Produto.nome)
                == nome.lower()
            )
            .first()
        )

        if produto_existente:
            flash(
                "Já existe um produto com esse nome.",
                "erro"
            )

            return redirect(
                url_for("produtos")
            )

        if preco_venda < 0:
            flash(
                "O preço de venda não pode ser negativo.",
                "erro"
            )

            return redirect(
                url_for("produtos")
            )

        if tipo_produto == "Revenda":
            if custo_compra < 0:
                flash(
                    "O custo de compra não pode ser negativo.",
                    "erro"
                )

                return redirect(
                    url_for("produtos")
                )

            if estoque_produto < 0:
                flash(
                    "O estoque inicial não pode ser negativo.",
                    "erro"
                )

                return redirect(
                    url_for("produtos")
                )

            rendimento_quantidade = 1
            rendimento_unidade = "un"

        else:
            tipo_produto = "Produzido"
            custo_compra = 0
            estoque_produto = 0

            if rendimento_quantidade <= 0:
                flash(
                    (
                        "O rendimento do produto produzido "
                        "deve ser maior que zero."
                    ),
                    "erro"
                )

                return redirect(
                    url_for("produtos")
                )

        if finalidade == "Preparo Interno":
            preco_venda = 0

        novo_produto = Produto(
            nome=nome,
            categoria=categoria,
            preco_venda=preco_venda,
            ativo=True,
            tipo_produto=tipo_produto,
            custo_compra=custo_compra,
            estoque_produto=estoque_produto,
            finalidade=finalidade,
            rendimento_quantidade=(
                rendimento_quantidade
            ),
            rendimento_unidade=(
                rendimento_unidade
            )
        )

        try:
            db.session.add(
                novo_produto
            )

            db.session.flush()

            # Registra o saldo inicial do produto de revenda.
            if (
                tipo_produto == "Revenda"
                and estoque_produto > 0
            ):
                valor_estoque_inicial = (
                    estoque_produto
                    * custo_compra
                )

                movimentacao_inicial = (
                    MovimentacaoProduto(
                        produto_id=novo_produto.id,
                        venda_id=None,
                        tipo="Entrada",
                        quantidade=estoque_produto,
                        valor_total=(
                            valor_estoque_inicial
                        ),
                        observacao=(
                            "Estoque inicial informado "
                            "no cadastro do produto."
                        ),
                        data=datetime.now()
                    )
                )

                db.session.add(
                    movimentacao_inicial
                )

            db.session.commit()

            flash(
                "Produto cadastrado com sucesso.",
                "sucesso"
            )

        except Exception as erro:
            db.session.rollback()

            print(
                "Erro ao cadastrar produto:",
                erro
            )

            flash(
                "Não foi possível cadastrar o produto.",
                "erro"
            )

        return redirect(
            url_for("produtos")
        )

    produtos_lista = (
        Produto.query
        .order_by(
            Produto.nome.asc()
        )
        .all()
    )

    return render_template(
        "produtos.html",
        produtos=produtos_lista
    )


# =========================================================
# EDITAR PRODUTO
# =========================================================

@app.route(
    "/editar_produto/<int:produto_id>",
    methods=["POST"]
)
def editar_produto(produto_id):
    if "usuario_id" not in session:
        return redirect(
            url_for("login")
        )

    produto = Produto.query.get_or_404(
        produto_id
    )

    nome = (
        request.form.get("nome", "")
        .strip()
    )

    categoria = (
        request.form.get(
            "categoria",
            "Outros"
        )
        .strip()
    )

    tipo_produto = (
        request.form.get(
            "tipo_produto",
            produto.tipo_produto
        )
        .strip()
    )

    finalidade = (
        request.form.get(
            "finalidade",
            produto.finalidade
        )
        .strip()
    )

    preco_venda = converter_numero_formulario(
        request.form.get("preco_venda"),
        produto.preco_venda or 0
    )

    custo_compra = converter_numero_formulario(
        request.form.get("custo_compra"),
        produto.custo_compra or 0
    )

    estoque_informado = converter_numero_formulario(
        request.form.get("estoque_produto"),
        produto.estoque_produto or 0
    )

    rendimento_quantidade = (
        converter_numero_formulario(
            request.form.get(
                "rendimento_quantidade"
            ),
            produto.rendimento_quantidade or 1
        )
    )

    rendimento_unidade = (
        request.form.get(
            "rendimento_unidade",
            produto.rendimento_unidade or "un"
        )
        .strip()
    )

    if not nome:
        flash(
            "Informe o nome do produto.",
            "erro"
        )

        return redirect(
            url_for("produtos")
        )

    produto_com_mesmo_nome = (
        Produto.query
        .filter(
            db.func.lower(Produto.nome)
            == nome.lower(),
            Produto.id != produto.id
        )
        .first()
    )

    if produto_com_mesmo_nome:
        flash(
            "Já existe outro produto com esse nome.",
            "erro"
        )

        return redirect(
            url_for("produtos")
        )

    if preco_venda < 0:
        flash(
            "O preço de venda não pode ser negativo.",
            "erro"
        )

        return redirect(
            url_for("produtos")
        )

    estoque_anterior = float(
        produto.estoque_produto or 0
    )

    tipo_anterior = produto.tipo_produto

    if tipo_produto == "Revenda":
        if custo_compra < 0:
            flash(
                "O custo de compra não pode ser negativo.",
                "erro"
            )

            return redirect(
                url_for("produtos")
            )

        if estoque_informado < 0:
            flash(
                "O estoque não pode ser negativo.",
                "erro"
            )

            return redirect(
                url_for("produtos")
            )

        rendimento_quantidade = 1
        rendimento_unidade = "un"

    else:
        tipo_produto = "Produzido"
        custo_compra = 0
        estoque_informado = 0

        if rendimento_quantidade <= 0:
            flash(
                "O rendimento deve ser maior que zero.",
                "erro"
            )

            return redirect(
                url_for("produtos")
            )

    if finalidade == "Preparo Interno":
        preco_venda = 0

    try:
        produto.nome = nome
        produto.categoria = categoria
        produto.preco_venda = preco_venda
        produto.tipo_produto = tipo_produto
        produto.finalidade = finalidade
        produto.rendimento_quantidade = (
            rendimento_quantidade
        )
        produto.rendimento_unidade = (
            rendimento_unidade
        )

        if tipo_produto == "Revenda":
            produto.custo_compra = custo_compra
            produto.estoque_produto = (
                estoque_informado
            )

            diferenca_estoque = (
                estoque_informado
                - estoque_anterior
            )

            if (
                tipo_anterior == "Revenda"
                and abs(diferenca_estoque) > 0.000001
            ):
                valor_ajuste = (
                    abs(diferenca_estoque)
                    * custo_compra
                )

                observacao = (
                    "Ajuste manual de entrada "
                    "realizado na edição do produto."
                    if diferenca_estoque > 0
                    else
                    "Ajuste manual de saída "
                    "realizado na edição do produto."
                )

                movimentacao_ajuste = (
                    MovimentacaoProduto(
                        produto_id=produto.id,
                        venda_id=None,
                        tipo=(
                            "Entrada"
                            if diferenca_estoque > 0
                            else "Saída"
                        ),
                        quantidade=abs(
                            diferenca_estoque
                        ),
                        valor_total=valor_ajuste,
                        observacao=observacao,
                        data=datetime.now()
                    )
                )

                db.session.add(
                    movimentacao_ajuste
                )

            elif (
                tipo_anterior != "Revenda"
                and estoque_informado > 0
            ):
                movimentacao_conversao = (
                    MovimentacaoProduto(
                        produto_id=produto.id,
                        venda_id=None,
                        tipo="Entrada",
                        quantidade=estoque_informado,
                        valor_total=(
                            estoque_informado
                            * custo_compra
                        ),
                        observacao=(
                            "Estoque inicial registrado "
                            "após alteração para revenda."
                        ),
                        data=datetime.now()
                    )
                )

                db.session.add(
                    movimentacao_conversao
                )

        else:
            produto.custo_compra = 0
            produto.estoque_produto = 0

        db.session.commit()

        flash(
            "Produto atualizado com sucesso.",
            "sucesso"
        )

    except Exception as erro:
        db.session.rollback()

        print(
            "Erro ao editar produto:",
            erro
        )

        flash(
            "Não foi possível atualizar o produto.",
            "erro"
        )

    return redirect(
        url_for("produtos")
    )


# =========================================================
# REGISTRAR COMPRA DE PRODUTO DE REVENDA
# =========================================================

@app.route(
    "/registrar_compra_produto/<int:produto_id>",
    methods=["POST"]
)
def registrar_compra_produto(produto_id):
    if "usuario_id" not in session:
        return redirect(
            url_for("login")
        )

    produto = Produto.query.get_or_404(
        produto_id
    )

    if produto.tipo_produto != "Revenda":
        flash(
            (
                "A entrada por compra está disponível "
                "somente para produtos de revenda."
            ),
            "erro"
        )

        return redirect(
            url_for("produtos")
        )

    quantidade_comprada = (
        converter_numero_formulario(
            request.form.get("quantidade"),
            0
        )
    )

    valor_total_compra = (
        converter_numero_formulario(
            request.form.get("valor_total"),
            0
        )
    )

    data_texto = (
        request.form.get("data", "")
        .strip()
    )

    observacao = (
        request.form.get(
            "observacao",
            ""
        )
        .strip()
    )

    if quantidade_comprada <= 0:
        flash(
            (
                "A quantidade comprada deve ser "
                "maior que zero."
            ),
            "erro"
        )

        return redirect(
            url_for("produtos")
        )

    if valor_total_compra <= 0:
        flash(
            (
                "O valor total da compra deve ser "
                "maior que zero."
            ),
            "erro"
        )

        return redirect(
            url_for("produtos")
        )

    try:
        if data_texto:
            data_compra = datetime.strptime(
                data_texto,
                "%Y-%m-%d"
            )

        else:
            data_compra = datetime.now()

    except ValueError:
        flash(
            "A data da compra é inválida.",
            "erro"
        )

        return redirect(
            url_for("produtos")
        )

    estoque_atual = float(
        produto.estoque_produto or 0
    )

    custo_medio_atual = float(
        produto.custo_compra or 0
    )

    valor_estoque_atual = (
        estoque_atual
        * custo_medio_atual
    )

    novo_estoque = (
        estoque_atual
        + quantidade_comprada
    )

    novo_valor_estoque = (
        valor_estoque_atual
        + valor_total_compra
    )

    if novo_estoque <= 0:
        novo_custo_medio = 0

    else:
        novo_custo_medio = (
            novo_valor_estoque
            / novo_estoque
        )

    try:
        produto.estoque_produto = (
            novo_estoque
        )

        produto.custo_compra = (
            novo_custo_medio
        )

        movimentacao = MovimentacaoProduto(
            produto_id=produto.id,
            venda_id=None,
            tipo="Entrada",
            quantidade=quantidade_comprada,
            valor_total=valor_total_compra,
            observacao=(
                observacao
                or "Compra de produto de revenda."
            ),
            data=data_compra
        )

        db.session.add(
            movimentacao
        )

        db.session.commit()

        custo_unitario_compra = (
            valor_total_compra
            / quantidade_comprada
        )

        flash(
            (
                f"Compra registrada com sucesso. "
                f"Custo unitário da compra: "
                f"R$ {custo_unitario_compra:.2f}. "
                f"Novo custo médio: "
                f"R$ {novo_custo_medio:.2f}."
            ),
            "sucesso"
        )

    except Exception as erro:
        db.session.rollback()

        print(
            "Erro ao registrar compra:",
            erro
        )

        flash(
            "Não foi possível registrar a compra.",
            "erro"
        )

    return redirect(
        url_for("produtos")
    )


# =========================================================
# ALTERAR STATUS DO PRODUTO
# =========================================================

@app.route(
    "/alterar_status_produto/<int:produto_id>",
    methods=["POST"]
)
def alterar_status_produto(produto_id):
    if "usuario_id" not in session:
        return redirect(
            url_for("login")
        )

    produto = Produto.query.get_or_404(
        produto_id
    )

    try:
        produto.ativo = not produto.ativo

        db.session.commit()

        status = (
            "ativado"
            if produto.ativo
            else "desativado"
        )

        flash(
            f"Produto {status} com sucesso.",
            "sucesso"
        )

    except Exception as erro:
        db.session.rollback()

        print(
            "Erro ao alterar status do produto:",
            erro
        )

        flash(
            (
                "Não foi possível alterar "
                "o status do produto."
            ),
            "erro"
        )

    return redirect(
        url_for("produtos")
    )


# =========================================================
# EXCLUIR PRODUTO
# =========================================================

@app.route(
    "/excluir_produto/<int:produto_id>",
    methods=["GET", "POST"]
)
def excluir_produto(produto_id):
    if "usuario_id" not in session:
        return redirect(
            url_for("login")
        )

    produto = Produto.query.get_or_404(
        produto_id
    )

    possui_vendas = (
        Venda.query
        .filter_by(
            produto_id=produto.id
        )
        .first()
    )

    if possui_vendas:
        flash(
            (
                "O produto não pode ser excluído porque "
                "possui vendas registradas. Desative-o "
                "para preservar o histórico."
            ),
            "erro"
        )

        return redirect(
            url_for("produtos")
        )

    utilizado_em_ficha = (
        FichaTecnica.query
        .filter_by(
            produto_base_id=produto.id
        )
        .first()
    )

    if utilizado_em_ficha:
        flash(
            (
                "O produto não pode ser excluído porque "
                "é utilizado na ficha técnica de outro "
                "produto."
            ),
            "erro"
        )

        return redirect(
            url_for("produtos")
        )

    try:
        db.session.delete(
            produto
        )

        db.session.commit()

        flash(
            "Produto excluído com sucesso.",
            "sucesso"
        )

    except Exception as erro:
        db.session.rollback()

        print(
            "Erro ao excluir produto:",
            erro
        )

        flash(
            "Não foi possível excluir o produto.",
            "erro"
        )

    return redirect(
        url_for("produtos")
    )


@app.route("/ficha_tecnica", methods=["GET", "POST"])
def ficha_tecnica():
    if not usuario_logado():
        return redirect(url_for("login"))

    produto_selecionado_id = request.args.get(
        "produto_id",
        type=int
    )

    if request.method == "POST":
        produto_id = request.form.get(
            "produto_id",
            type=int
        )

        tipo_item = request.form.get(
            "tipo_item",
            ""
        ).strip()

        quantidade = converter_float(
            request.form.get("quantidade")
        )

        unidade_utilizada = request.form.get(
            "unidade_utilizada",
            ""
        ).strip()

        # =========================
        # VALIDAÇÃO DO PRODUTO
        # =========================

        if not produto_id:
            flash(
                "Selecione o produto da ficha técnica.",
                "erro"
            )
            return redirect(
                url_for("ficha_tecnica")
            )

        produto = Produto.query.get_or_404(
            produto_id
        )

        # =========================
        # VALIDAÇÃO DA QUANTIDADE
        # =========================

        if quantidade <= 0:
            flash(
                "A quantidade deve ser maior que zero.",
                "erro"
            )
            return redirect(
                url_for(
                    "ficha_tecnica",
                    produto_id=produto_id
                )
            )

        # =========================
        # VALIDAÇÃO DA UNIDADE
        # =========================

        unidades_permitidas = {
            "g",
            "kg",
            "ml",
            "L",
            "un"
        }

        if unidade_utilizada not in unidades_permitidas:
            flash(
                "Selecione uma unidade de medida válida.",
                "erro"
            )
            return redirect(
                url_for(
                    "ficha_tecnica",
                    produto_id=produto_id
                )
            )

        # =========================
        # NOVO ITEM DA FICHA
        # =========================

        item = FichaTecnica(
            produto_id=produto_id,
            quantidade=quantidade,
            unidade_utilizada=unidade_utilizada
        )

        # =========================
        # COMPONENTE DO TIPO INSUMO
        # =========================

        if tipo_item == "insumo":
            insumo_id = request.form.get(
                "insumo_id",
                type=int
            )

            if not insumo_id:
                flash(
                    "Selecione um insumo.",
                    "erro"
                )
                return redirect(
                    url_for(
                        "ficha_tecnica",
                        produto_id=produto_id
                    )
                )

            insumo = Insumo.query.get_or_404(
                insumo_id
            )

            # Impede o mesmo insumo de ser
            # adicionado duas vezes à ficha
            item_existente = FichaTecnica.query.filter_by(
                produto_id=produto_id,
                insumo_id=insumo_id
            ).first()

            if item_existente:
                flash(
                    f"O insumo '{insumo.nome}' já está "
                    f"cadastrado na ficha de "
                    f"'{produto.nome}'.",
                    "erro"
                )
                return redirect(
                    url_for(
                        "ficha_tecnica",
                        produto_id=produto_id
                    )
                )

            item.insumo_id = insumo_id
            item.produto_base_id = None

        # =============================
        # COMPONENTE: PREPARO INTERNO
        # =============================

        elif tipo_item == "produto":
            produto_base_id = request.form.get(
                "produto_base_id",
                type=int
            )

            if not produto_base_id:
                flash(
                    "Selecione um preparo interno.",
                    "erro"
                )
                return redirect(
                    url_for(
                        "ficha_tecnica",
                        produto_id=produto_id
                    )
                )

            produto_base = Produto.query.get_or_404(
                produto_base_id
            )

            # Impede o produto de usar ele mesmo
            if produto_base.id == produto.id:
                flash(
                    "Um produto não pode utilizar a si "
                    "mesmo como componente.",
                    "erro"
                )
                return redirect(
                    url_for(
                        "ficha_tecnica",
                        produto_id=produto_id
                    )
                )

            # Somente preparos internos produzidos
            # podem ser usados como produto-base
            if (
                produto_base.tipo_produto != "Produzido"
                or produto_base.finalidade
                != "Preparo Interno"
            ):
                flash(
                    "Somente preparos internos produzidos "
                    "pela empresa podem ser utilizados "
                    "como base.",
                    "erro"
                )
                return redirect(
                    url_for(
                        "ficha_tecnica",
                        produto_id=produto_id
                    )
                )

            # Impede o mesmo preparo interno de ser
            # adicionado duas vezes
            item_existente = FichaTecnica.query.filter_by(
                produto_id=produto_id,
                produto_base_id=produto_base_id
            ).first()

            if item_existente:
                flash(
                    f"O preparo interno "
                    f"'{produto_base.nome}' já está "
                    f"cadastrado na ficha de "
                    f"'{produto.nome}'.",
                    "erro"
                )
                return redirect(
                    url_for(
                        "ficha_tecnica",
                        produto_id=produto_id
                    )
                )

            # Impede referências circulares
            if ficha_cria_ciclo(
                produto_id,
                produto_base_id
            ):
                flash(
                    "Essa inclusão criaria uma "
                    "referência circular entre as "
                    "fichas técnicas.",
                    "erro"
                )
                return redirect(
                    url_for(
                        "ficha_tecnica",
                        produto_id=produto_id
                    )
                )

            item.insumo_id = None
            item.produto_base_id = produto_base_id

        else:
            flash(
                "Selecione se o componente é um "
                "insumo ou um preparo interno.",
                "erro"
            )
            return redirect(
                url_for(
                    "ficha_tecnica",
                    produto_id=produto_id
                )
            )

        # =========================
        # SALVAMENTO
        # =========================

        try:
            db.session.add(item)
            db.session.commit()

            flash(
                "Componente adicionado à ficha "
                "técnica com sucesso!",
                "sucesso"
            )

        except Exception:
            db.session.rollback()

            flash(
                "Não foi possível adicionar o "
                "componente à ficha técnica.",
                "erro"
            )

        return redirect(
            url_for(
                "ficha_tecnica",
                produto_id=produto_id
            )
        )

    # =========================
    # LISTA DE PRODUTOS
    # =========================

    produtos_lista = Produto.query.filter_by(
        ativo=True
    ).order_by(
        Produto.nome
    ).all()

    # =========================
    # PREPAROS INTERNOS
    # =========================

    produtos_base = Produto.query.filter_by(
        tipo_produto="Produzido",
        finalidade="Preparo Interno",
        ativo=True
    ).order_by(
        Produto.nome
    ).all()

    # =========================
    # LISTA DE INSUMOS
    # =========================

    insumos_lista = Insumo.query.order_by(
        Insumo.nome
    ).all()

    produto_selecionado = None
    itens = []

    # Mostra somente a ficha do
    # produto selecionado
    if produto_selecionado_id:
        produto_selecionado = (
            Produto.query.get_or_404(
                produto_selecionado_id
            )
        )

        itens = FichaTecnica.query.filter_by(
            produto_id=produto_selecionado_id
        ).order_by(
            FichaTecnica.id
        ).all()

    return render_template(
        "ficha_tecnica.html",
        produtos=produtos_lista,
        produtos_base=produtos_base,
        insumos=insumos_lista,
        produto_selecionado=produto_selecionado,
        produto_selecionado_id=produto_selecionado_id,
        itens=itens
    )


@app.route(
    "/excluir_item_ficha/<int:id>",
    methods=["POST"]
)
def excluir_item_ficha(id):
    if not usuario_logado():
        return redirect(url_for("login"))

    item = FichaTecnica.query.get_or_404(id)

    produto_id = item.produto_id

    try:
        db.session.delete(item)
        db.session.commit()

        flash(
            "Componente removido da ficha técnica.",
            "sucesso"
        )

    except Exception:
        db.session.rollback()

        flash(
            "Não foi possível remover o componente "
            "da ficha técnica.",
            "erro"
        )

    return redirect(
        url_for(
            "ficha_tecnica",
            produto_id=produto_id
        )
    )


# =========================================================
# VENDAS
# =========================================================

@app.route(
    "/vendas",
    methods=["GET", "POST"],
)
def vendas():
    if not usuario_logado():
        return redirect(
            url_for("login")
        )

    if request.method == "POST":
        produto_id = request.form.get(
            "produto_id"
        )

        quantidade = converter_float(
            request.form.get("quantidade"),
            0,
        )

        data_venda_texto = (
            request.form.get(
                "data_venda",
                "",
            )
            or request.form.get(
                "data",
                "",
            )
            or ""
        ).strip()

        movimentar_estoque_texto = (
            request.form.get(
                "movimentar_estoque",
                "sim",
            )
            or "sim"
        ).strip().lower()

        movimentar_estoque = (
            movimentar_estoque_texto
            not in {
                "nao",
                "não",
                "false",
                "0",
            }
        )

        # -------------------------------------------------
        # VALIDAÇÕES INICIAIS
        # -------------------------------------------------

        if not produto_id:
            flash(
                "Selecione um produto.",
                "erro",
            )

            return redirect(
                url_for("vendas")
            )

        try:
            produto_id = int(
                produto_id
            )

        except (TypeError, ValueError):
            flash(
                "O produto informado é inválido.",
                "erro",
            )

            return redirect(
                url_for("vendas")
            )

        if quantidade <= 0:
            flash(
                "A quantidade vendida deve ser maior que zero.",
                "erro",
            )

            return redirect(
                url_for("vendas")
            )

        produto = Produto.query.get_or_404(
            produto_id
        )

        if not produto.ativo:
            flash(
                "Não é possível registrar a venda de um produto desativado.",
                "erro",
            )

            return redirect(
                url_for("vendas")
            )

        if produto.finalidade != "Venda":
            flash(
                "O item selecionado é um preparo interno e não pode ser vendido diretamente.",
                "erro",
            )

            return redirect(
                url_for("vendas")
            )

        # -------------------------------------------------
        # DATA DA VENDA
        # -------------------------------------------------

        if data_venda_texto:
            try:
                data_venda = datetime.strptime(
                    data_venda_texto,
                    "%Y-%m-%d",
                )

            except ValueError:
                flash(
                    "A data informada para a venda é inválida.",
                    "erro",
                )

                return redirect(
                    url_for("vendas")
                )

        else:
            data_venda = datetime.now()

        # -------------------------------------------------
        # CÁLCULOS DA VENDA
        # -------------------------------------------------

        preco_unitario = float(
            produto.preco_venda or 0
        )

        custo_unitario = float(
            produto.custo_materia_prima()
            or 0
        )

        receita_total = (
            preco_unitario
            * quantidade
        )

        cmv_total = (
            custo_unitario
            * quantidade
        )

        margem_total = (
            receita_total
            - cmv_total
        )

        # -------------------------------------------------
        # VALIDAÇÃO DO ESTOQUE DE REVENDA
        # -------------------------------------------------

        if (
            movimentar_estoque
            and produto.tipo_produto == "Revenda"
        ):
            estoque_atual = float(
                produto.estoque_produto or 0
            )

            if estoque_atual < quantidade:
                flash(
                    (
                        f"Estoque insuficiente para {produto.nome}. "
                        f"Disponível: {estoque_atual:.2f}. "
                        f"Quantidade solicitada: {quantidade:.2f}."
                    ),
                    "erro",
                )

                return redirect(
                    url_for("vendas")
                )

        # -------------------------------------------------
        # VALIDAÇÃO DOS INSUMOS DO PRODUTO PRODUZIDO
        # -------------------------------------------------

        if (
            movimentar_estoque
            and produto.tipo_produto == "Produzido"
        ):
            for item in produto.ficha_itens:
                if not item.insumo:
                    continue

                quantidade_por_unidade = float(
                    item.quantidade_convertida_para_estoque()
                    or 0
                )

                quantidade_necessaria = (
                    quantidade_por_unidade
                    * quantidade
                )

                estoque_disponivel = float(
                    item.insumo.estoque_atual()
                    or 0
                )

                if (
                    estoque_disponivel
                    < quantidade_necessaria
                ):
                    flash(
                        (
                            f"Estoque insuficiente de "
                            f"{item.insumo.nome}. "
                            f"Disponível: "
                            f"{estoque_disponivel:.3f} "
                            f"{item.insumo.unidade}. "
                            f"Necessário: "
                            f"{quantidade_necessaria:.3f} "
                            f"{item.insumo.unidade}."
                        ),
                        "erro",
                    )

                    return redirect(
                        url_for("vendas")
                    )

        try:
            # ---------------------------------------------
            # CRIA A VENDA
            # ---------------------------------------------

            nova_venda = Venda(
                produto_id=produto.id,
                data=data_venda,
                quantidade=quantidade,
                receita_total=receita_total,
                cmv_total=cmv_total,
                margem_total=margem_total,
                movimentou_estoque=movimentar_estoque,
            )

            db.session.add(
                nova_venda
            )

            # Gera o ID da venda antes das movimentações.
            db.session.flush()

            # ---------------------------------------------
            # BAIXA DO PRODUTO DE REVENDA
            # ---------------------------------------------

            if (
                movimentar_estoque
                and produto.tipo_produto == "Revenda"
            ):
                valor_saida = (
                    quantidade
                    * float(
                        produto.custo_compra or 0
                    )
                )

                produto.estoque_produto = (
                    float(
                        produto.estoque_produto or 0
                    )
                    - quantidade
                )

                movimentacao_produto = (
                    MovimentacaoProduto(
                        produto_id=produto.id,
                        venda_id=nova_venda.id,
                        tipo="Saída",
                        quantidade=quantidade,
                        valor_total=valor_saida,
                        observacao=(
                            f"Saída referente à venda "
                            f"nº {nova_venda.id}."
                        ),
                        data=data_venda,
                    )
                )

                db.session.add(
                    movimentacao_produto
                )

            # ---------------------------------------------
            # BAIXA DOS INSUMOS DO PRODUTO PRODUZIDO
            # ---------------------------------------------

            elif (
                movimentar_estoque
                and produto.tipo_produto == "Produzido"
            ):
                for item in produto.ficha_itens:
                    if not item.insumo:
                        continue

                    quantidade_por_unidade = float(
                        item.quantidade_convertida_para_estoque()
                        or 0
                    )

                    quantidade_saida = (
                        quantidade_por_unidade
                        * quantidade
                    )

                    custo_medio_insumo = float(
                        item.insumo.custo_medio_unitario()
                        or 0
                    )

                    valor_saida = (
                        quantidade_saida
                        * custo_medio_insumo
                    )

                    movimentacao_insumo = (
                        MovimentacaoEstoque(
                            insumo_id=item.insumo.id,
                            venda_id=nova_venda.id,
                            tipo="Saída",
                            quantidade=quantidade_saida,
                            valor_total=valor_saida,
                            observacao=(
                                f"Saída referente à venda "
                                f"nº {nova_venda.id} — "
                                f"{produto.nome}."
                            ),
                            data=data_venda,
                        )
                    )

                    db.session.add(
                        movimentacao_insumo
                    )

            db.session.commit()

            if movimentar_estoque:
                mensagem = (
                    "Venda registrada com sucesso e estoque atualizado."
                )

            else:
                mensagem = (
                    "Venda retroativa registrada sem movimentar o estoque."
                )

            flash(
                mensagem,
                "sucesso",
            )

        except Exception as erro:
            db.session.rollback()

            print(
                "Erro ao registrar venda:",
                erro,
            )

            flash(
                (
                    "Não foi possível registrar a venda. "
                    "Nenhuma alteração foi realizada."
                ),
                "erro",
            )

        return redirect(
            url_for("vendas")
        )

    # =====================================================
    # ABERTURA DA PÁGINA DE VENDAS
    # =====================================================

    produtos_lista = (
        Produto.query
        .filter_by(
            ativo=True,
            finalidade="Venda",
        )
        .order_by(
            Produto.nome
        )
        .all()
    )

    vendas_lista = (
        Venda.query
        .order_by(
            Venda.data.desc(),
            Venda.id.desc(),
        )
        .all()
    )

    receita_total = sum(
        float(
            venda.receita_total or 0
        )
        for venda in vendas_lista
    )

    cmv_total = sum(
        float(
            venda.cmv_total or 0
        )
        for venda in vendas_lista
    )

    margem_total = sum(
        float(
            venda.margem_total or 0
        )
        for venda in vendas_lista
    )

    quantidade_total = sum(
        float(
            venda.quantidade or 0
        )
        for venda in vendas_lista
    )

    vendas_retroativas = sum(
        1
        for venda in vendas_lista
        if not venda.movimentou_estoque
    )

    return render_template(
        "vendas.html",
        produtos=produtos_lista,
        vendas=vendas_lista,
        receita_total=receita_total,
        cmv_total=cmv_total,
        margem_total=margem_total,
        quantidade_total=quantidade_total,
        vendas_retroativas=vendas_retroativas,
    )


# =========================================================
# EXCLUIR VENDA
# =========================================================

@app.route(
    "/excluir_venda/<int:venda_id>",
    methods=["POST", "GET"],
)
def excluir_venda(venda_id):
    if not usuario_logado():
        return redirect(
            url_for("login")
        )

    venda = Venda.query.get_or_404(
        venda_id
    )

    produto = venda.produto

    try:
        # -------------------------------------------------
        # VENDA RETROATIVA
        # -------------------------------------------------

        if not venda.movimentou_estoque:
            db.session.delete(
                venda
            )

            db.session.commit()

            flash(
                "Venda retroativa excluída com sucesso.",
                "sucesso",
            )

            return redirect(
                url_for("vendas")
            )

        # -------------------------------------------------
        # DEVOLUÇÃO DE PRODUTO DE REVENDA
        # -------------------------------------------------

        if (
            produto
            and produto.tipo_produto == "Revenda"
        ):
            movimentacoes_produto = (
                MovimentacaoProduto.query
                .filter_by(
                    venda_id=venda.id,
                    tipo="Saída",
                )
                .all()
            )

            quantidade_devolver = sum(
                float(
                    movimentacao.quantidade or 0
                )
                for movimentacao
                in movimentacoes_produto
            )

            # Compatibilidade com vendas antigas que ainda
            # não possuem MovimentacaoProduto vinculada.
            if quantidade_devolver <= 0:
                quantidade_devolver = float(
                    venda.quantidade or 0
                )

            produto.estoque_produto = (
                float(
                    produto.estoque_produto or 0
                )
                + quantidade_devolver
            )

            for movimentacao in movimentacoes_produto:
                db.session.delete(
                    movimentacao
                )

        # -------------------------------------------------
        # DEVOLUÇÃO DOS INSUMOS DO PRODUTO PRODUZIDO
        # -------------------------------------------------

        elif (
            produto
            and produto.tipo_produto == "Produzido"
        ):
            movimentacoes_insumo = (
                MovimentacaoEstoque.query
                .filter_by(
                    venda_id=venda.id,
                    tipo="Saída",
                )
                .all()
            )

            for movimentacao in movimentacoes_insumo:
                db.session.delete(
                    movimentacao
                )

        # As movimentações de saída são removidas.
        # Com isso, o estoque calculado dos insumos volta
        # automaticamente ao saldo anterior.

        db.session.delete(
            venda
        )

        db.session.commit()

        flash(
            (
                "Venda excluída com sucesso e "
                "estoque restaurado."
            ),
            "sucesso",
        )

    except Exception as erro:
        db.session.rollback()

        print(
            "Erro ao excluir venda:",
            erro,
        )

        flash(
            (
                "Não foi possível excluir a venda. "
                "Nenhuma alteração foi realizada."
            ),
            "erro",
        )

    return redirect(
        url_for("vendas")
    )

@app.route("/financeiro", methods=["GET", "POST"])
def financeiro():
    if not usuario_logado():
        return redirect(url_for("login"))

    if request.method == "POST":
        valor = converter_float(
            request.form.get("valor")
        )

        if valor <= 0:
            flash("O valor deve ser maior que zero.")
            return redirect(url_for("financeiro"))

        novo = Financeiro(
            tipo=request.form.get("tipo", "Saída"),
            categoria=request.form.get(
                "categoria",
                "",
            ).strip(),
            descricao=request.form.get(
                "descricao",
                "",
            ).strip(),
            valor=valor,
        )

        db.session.add(novo)
        db.session.commit()

        flash("Lançamento financeiro registrado!")
        return redirect(url_for("financeiro"))

    registros = Financeiro.query.order_by(
        Financeiro.data.desc()
    ).all()

    total_entradas = sum(
        registro.valor
        for registro in registros
        if registro.tipo == "Entrada"
    )

    total_saidas = sum(
        registro.valor
        for registro in registros
        if registro.tipo == "Saída"
    )

    saldo = total_entradas - total_saidas

    return render_template(
        "financeiro.html",
        registros=registros,
        total_entradas=total_entradas,
        total_saidas=total_saidas,
        saldo=saldo,
    )


@app.route("/excluir_financeiro/<int:id>", methods=["POST", "GET"])
def excluir_financeiro(id):
    if not usuario_logado():
        return redirect(url_for("login"))

    registro = Financeiro.query.get_or_404(id)

    db.session.delete(registro)
    db.session.commit()

    flash("Lançamento financeiro excluído!")
    return redirect(url_for("financeiro"))


@app.route("/relatorios")
def relatorios():
    if not usuario_logado():
        return redirect(url_for("login"))

    insumos_lista = Insumo.query.order_by(
        Insumo.nome
    ).all()

    produtos_lista = Produto.query.order_by(
        Produto.nome
    ).all()

    vendas_lista = Venda.query.order_by(
        Venda.data.desc()
    ).all()

    financeiros = Financeiro.query.order_by(
        Financeiro.data.desc()
    ).all()

    receita_total = sum(
        venda.receita_total or 0
        for venda in vendas_lista
    )

    cmv_total = sum(
        venda.cmv_total or 0
        for venda in vendas_lista
    )

    margem_total = sum(
        venda.margem_total or 0
        for venda in vendas_lista
    )

    despesas_operacionais = sum(
        financeiro.valor or 0
        for financeiro in financeiros
        if financeiro.tipo == "Saída"
    )

    lucro_operacional = (
        margem_total
        - despesas_operacionais
    )

    margem_percentual = 0

    if receita_total > 0:
        margem_percentual = (
            margem_total / receita_total
        ) * 100

    valor_estoque = sum(
        insumo.estoque_atual()
        * insumo.custo_medio_unitario()
        for insumo in insumos_lista
    )

    valor_estoque_produtos = sum(
        (produto.estoque_produto or 0)
        * (
            produto.custo_compra
            if produto.tipo_produto == "Revenda"
            else produto.custo_materia_prima()
        )
        for produto in produtos_lista
    )

    valor_estoque += valor_estoque_produtos

    total_vendido = sum(
        venda.quantidade or 0
        for venda in vendas_lista
    )

    itens_abaixo_minimo = sum(
        1
        for insumo in insumos_lista
        if insumo.estoque_atual()
        <= insumo.estoque_minimo()
    )

    itens_ponto_pedido = sum(
        1
        for insumo in insumos_lista
        if insumo.estoque_minimo()
        < insumo.estoque_atual()
        <= insumo.ponto_pedido()
    )

    coberturas = []

    for insumo in insumos_lista:
        cobertura = insumo.cobertura_estoque()

        if cobertura > 0:
            coberturas.append(cobertura)

    cobertura_media = (
        sum(coberturas) / len(coberturas)
        if coberturas
        else 0
    )

    giros = []

    for insumo in insumos_lista:
        giro = insumo.giro_estoque()

        if giro > 0:
            giros.append(giro)

    giro_medio = (
        sum(giros) / len(giros)
        if giros
        else 0
    )

    lotes = []

    for insumo in insumos_lista:
        lote = insumo.lote_economico()

        if lote > 0:
            lotes.append(lote)

    lote_economico_medio = (
        sum(lotes) / len(lotes)
        if lotes
        else 0
    )

    ranking_produtos = {}

    for venda in vendas_lista:
        if venda.produto is None:
            continue

        nome_produto = venda.produto.nome

        if nome_produto not in ranking_produtos:
            ranking_produtos[nome_produto] = {
                "quantidade": 0,
                "receita": 0,
                "cmv": 0,
                "margem": 0,
            }

        ranking_produtos[nome_produto]["quantidade"] += (
            venda.quantidade or 0
        )

        ranking_produtos[nome_produto]["receita"] += (
            venda.receita_total or 0
        )

        ranking_produtos[nome_produto]["cmv"] += (
            venda.cmv_total or 0
        )

        ranking_produtos[nome_produto]["margem"] += (
            venda.margem_total or 0
        )

    ranking_produtos = sorted(
        ranking_produtos.items(),
        key=lambda item: item[1]["quantidade"],
        reverse=True,
    )

    return render_template(
        "relatorios.html",
        insumos=insumos_lista,
        produtos=produtos_lista,
        vendas=vendas_lista,
        financeiros=financeiros,
        receita_total=receita_total,
        cmv_total=cmv_total,
        margem_total=margem_total,
        margem_percentual=margem_percentual,
        despesas_operacionais=despesas_operacionais,
        lucro_operacional=lucro_operacional,
        valor_estoque=valor_estoque,
        total_vendido=total_vendido,
        ranking_produtos=ranking_produtos,
        itens_abaixo_minimo=itens_abaixo_minimo,
        itens_ponto_pedido=itens_ponto_pedido,
        cobertura_media=cobertura_media,
        giro_medio=giro_medio,
        lote_economico_medio=lote_economico_medio,
    )


@app.route("/exportar_caixa")
def exportar_caixa():
    if not usuario_logado():
        return redirect(url_for("login"))

    vendas_lista = Venda.query.order_by(
        Venda.data.asc()
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Caixa"

    ws["A1"] = "ERP RESTAURANTE"
    ws["A2"] = "FECHAMENTO DE CAIXA"
    ws["A3"] = (
        "Data de emissão: "
        f"{datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )

    ws.append([])

    ws.append(
        [
            "Data",
            "Produto",
            "Quantidade",
            "Receita",
            "CMV",
            "Margem de Contribuição",
        ]
    )

    receita_total = 0
    cmv_total = 0
    margem_total = 0
    quantidade_total = 0

    for venda in vendas_lista:
        nome_produto = (
            venda.produto.nome
            if venda.produto
            else "Produto removido"
        )

        ws.append(
            [
                venda.data.strftime("%d/%m/%Y %H:%M"),
                nome_produto,
                venda.quantidade,
                venda.receita_total,
                venda.cmv_total,
                venda.margem_total,
            ]
        )

        receita_total += venda.receita_total or 0
        cmv_total += venda.cmv_total or 0
        margem_total += venda.margem_total or 0
        quantidade_total += venda.quantidade or 0

    ws.append([])
    ws.append(["RESUMO DO CAIXA"])
    ws.append(["Quantidade vendida", quantidade_total])
    ws.append(["Receita total", receita_total])
    ws.append(["CMV total", cmv_total])
    ws.append(["Margem de contribuição", margem_total])

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 25

    for linha in ws.iter_rows(
        min_row=6,
        min_col=4,
        max_col=6,
    ):
        for celula in linha:
            celula.number_format = 'R$ #,##0.00'

    arquivo = io.BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    nome_arquivo = (
        "Fechamento_Caixa_"
        f"{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    )

    return send_file(
        arquivo,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


@app.route("/configuracoes")
def configuracoes():
    if not usuario_logado():
        return redirect(url_for("login"))

    total_insumos = Insumo.query.count()
    total_produtos = Produto.query.count()
    total_vendas = Venda.query.count()
    total_lancamentos = Financeiro.query.count()

    return render_template(
        "configuracoes.html",
        nome=session.get("usuario_nome"),
        total_insumos=total_insumos,
        total_produtos=total_produtos,
        total_vendas=total_vendas,
        total_lancamentos=total_lancamentos,
    )


@app.route("/fazer_backup")
def fazer_backup():
    if not usuario_logado():
        return redirect(url_for("login"))

    uri_banco = app.config["SQLALCHEMY_DATABASE_URI"]

    if not uri_banco.startswith("sqlite:///"):
        flash(
            "O backup local automático está disponível apenas para SQLite."
        )
        return redirect(url_for("configuracoes"))

    caminho_relativo = uri_banco.replace(
        "sqlite:///",
        "",
        1,
    )

    if os.path.isabs(caminho_relativo):
        origem = caminho_relativo
    else:
        origem = os.path.join(
            app.instance_path,
            caminho_relativo,
        )

    pasta_backup = os.path.join(
        app.root_path,
        "backups",
    )

    os.makedirs(
        pasta_backup,
        exist_ok=True,
    )

    data_hora = datetime.now().strftime(
        "%Y-%m-%d_%H-%M-%S"
    )

    destino = os.path.join(
        pasta_backup,
        f"backup_{data_hora}.db",
    )

    if os.path.exists(origem):
        shutil.copy2(origem, destino)
        flash("Backup realizado com sucesso!")
    else:
        flash(
            f"Banco de dados não encontrado em: {origem}"
        )

    return redirect(url_for("configuracoes"))


@app.route("/alterar_senha", methods=["POST"])
def alterar_senha():
    if not usuario_logado():
        return redirect(url_for("login"))

    usuario = db.session.get(
        Usuario,
        session["usuario_id"],
    )

    if usuario is None:
        session.clear()
        flash("Sua sessão expirou. Entre novamente.")
        return redirect(url_for("login"))

    senha_atual = request.form.get(
        "senha_atual",
        "",
    )

    nova_senha = request.form.get(
        "nova_senha",
        "",
    )

    confirmar = request.form.get(
        "confirmar_senha",
        "",
    )

    if not usuario.verificar_senha(senha_atual):
        flash("Senha atual incorreta.")
        return redirect(url_for("configuracoes"))

    if len(nova_senha) < 6:
        flash("A nova senha deve ter pelo menos 6 caracteres.")
        return redirect(url_for("configuracoes"))

    if nova_senha != confirmar:
        flash("As novas senhas não coincidem.")
        return redirect(url_for("configuracoes"))

    usuario.set_senha(nova_senha)
    db.session.commit()

    flash("Senha alterada com sucesso!")
    return redirect(url_for("configuracoes"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


if __name__ == "__main__":
    app.run(debug=True)